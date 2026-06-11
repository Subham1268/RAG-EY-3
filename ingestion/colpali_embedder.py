"""
ingestion/colpali_embedder.py
──────────────────────────────
ColPali page-level visual embeddings for superior retrieval of
documents with mixed text, tables, charts, and diagrams.

ColPali (vidore/colpali-v1.2) embeds entire page images directly —
no OCR, no chunking, no information loss from text extraction.

This runs ALONGSIDE the existing chunk pipeline:
  - Existing pipeline  → fine-grained chunk retrieval + granular citations
  - ColPali pipeline   → page-level retrieval for visually rich pages

At query time, results from both are merged and deduplicated by page.

Hardware:
  - GPU (8GB+ VRAM): fast, ~1-2s per page
  - CPU: very slow (~30s per page), not recommended for large ingestion
  - Cloud: run on a GPU VM for ingestion, CPU is fine for query-time
    ColPali query embedding (single vector, fast even on CPU)

Install:
    pip install byaldi torch torchvision pillow
    # For GPU: pip install torch --index-url https://download.pytorch.org/whl/cu118
"""

from __future__ import annotations

import asyncio
import base64
import io
from dataclasses import dataclass
from pathlib import Path

from PIL import Image

# ColPali via byaldi
try:
    from byaldi import RAGMultiModalModel
    COLPALI_AVAILABLE = True
except ImportError:
    COLPALI_AVAILABLE = False
    print("⚠️  byaldi not installed. ColPali retrieval disabled. Run: pip install byaldi")


@dataclass
class ColPaliPage:
    """A single page ready for ColPali indexing."""
    page_image_b64: str         # Base64 PNG of the full page
    page_idx: int               # 0-indexed page number
    source_file: str
    doc_type: str               # pdf | pptx | docx | xlsx
    metadata: dict              # engagement_id, client, country, practice, year, etc.


class ColPaliEmbedder:
    """
    Wraps the ColPali model for page-level embedding.

    Usage:
        embedder = ColPaliEmbedder()
        pages = embedder.extract_pages("report.pdf", metadata={...})
        vectors = embedder.embed_pages(pages)
    """

    MODEL_NAME = "vidore/colpali-v1.2"

    def __init__(self) -> None:
        if not COLPALI_AVAILABLE:
            raise RuntimeError("byaldi not installed. Run: pip install byaldi")
        self._model: RAGMultiModalModel | None = None

    def _get_model(self) -> RAGMultiModalModel:
        """Lazy load ColPali model (downloads on first use ~5GB)."""
        if self._model is None:
            print(f"Loading ColPali model: {self.MODEL_NAME} (first load downloads ~5GB)...")
            self._model = RAGMultiModalModel.from_pretrained(self.MODEL_NAME)
            print("ColPali model loaded.")
        return self._model

    # ── Page extraction ────────────────────────────────────────────────────────

    def extract_pages(self, file_path: str | Path, metadata: dict) -> list[ColPaliPage]:
        """
        Render every page of a document to a PIL Image and return ColPaliPage list.
        Supports PDF, PPTX (via LibreOffice), DOCX (via LibreOffice).
        """
        path = Path(file_path)
        ext  = path.suffix.lower().lstrip(".")

        if ext == "pdf":
            return self._extract_pdf_pages(path, metadata)
        elif ext in ("pptx", "docx"):
            return self._extract_via_libreoffice(path, metadata, ext)
        elif ext == "xlsx":
            # XLSX: render each sheet as a table image (simpler approach)
            return self._extract_xlsx_pages(path, metadata)
        else:
            return []

    def _extract_pdf_pages(self, path: Path, metadata: dict) -> list[ColPaliPage]:
        import fitz
        pages: list[ColPaliPage] = []
        fitz_doc = fitz.open(str(path))
        for page_idx, page in enumerate(fitz_doc):
            # 150 DPI is enough for ColPali; higher = more memory
            mat = fitz.Matrix(150 / 72, 150 / 72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            png_bytes = pix.tobytes("png")
            img_b64 = base64.b64encode(png_bytes).decode()
            pages.append(ColPaliPage(
                page_image_b64=img_b64,
                page_idx=page_idx,
                source_file=str(path),
                doc_type="pdf",
                metadata=metadata,
            ))
        fitz_doc.close()
        return pages

    def _extract_via_libreoffice(
        self, path: Path, metadata: dict, doc_type: str
    ) -> list[ColPaliPage]:
        """Convert PPTX/DOCX to PDF via LibreOffice, then render pages."""
        import subprocess, tempfile, os
        pages: list[ColPaliPage] = []
        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                # Convert to PDF
                result = subprocess.run(
                    ["libreoffice", "--headless", "--convert-to", "pdf",
                     "--outdir", tmp_dir, str(path)],
                    capture_output=True, timeout=60
                )
                pdf_files = list(Path(tmp_dir).glob("*.pdf"))
                if not pdf_files:
                    print(f"LibreOffice conversion failed for {path.name}")
                    return []
                # Render the converted PDF
                pages = self._extract_pdf_pages(pdf_files[0], metadata)
                for p in pages:
                    p.source_file = str(path)   # point back to original
                    p.doc_type    = doc_type
        except Exception as e:
            print(f"LibreOffice render failed for {path.name}: {e}")
        return pages

    def _extract_xlsx_pages(self, path: Path, metadata: dict) -> list[ColPaliPage]:
        """Render each XLSX sheet as a simple table image using PIL."""
        from openpyxl import load_workbook
        pages: list[ColPaliPage] = []
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))[:50]  # first 50 rows
                if not rows:
                    continue
                # Create a simple text render of the sheet
                lines = [f"Sheet: {sheet_name}"]
                for row in rows:
                    lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                text = "\n".join(lines)

                # Render to image using PIL
                img = Image.new("RGB", (1200, max(400, len(lines) * 20 + 60)), "white")
                try:
                    from PIL import ImageDraw, ImageFont
                    draw = ImageDraw.Draw(img)
                    draw.text((10, 10), text[:3000], fill="black")
                except Exception:
                    pass

                buf = io.BytesIO()
                img.save(buf, format="PNG")
                img_b64 = base64.b64encode(buf.getvalue()).decode()
                pages.append(ColPaliPage(
                    page_image_b64=img_b64,
                    page_idx=sheet_idx,
                    source_file=str(path),
                    doc_type="xlsx",
                    metadata={**metadata, "sheet_name": sheet_name},
                ))
            wb.close()
        except Exception as e:
            print(f"XLSX render failed for {path.name}: {e}")
        return pages

    # ── Embedding ──────────────────────────────────────────────────────────────

    def embed_pages(self, pages: list[ColPaliPage]) -> list[tuple[ColPaliPage, list[float]]]:
        """
        Embed a list of ColPaliPages using the ColPali model.
        Returns list of (page, embedding_vector) tuples.

        Note: ColPali produces multi-vector embeddings (one vector per patch).
        For Pinecone compatibility we average-pool to a single vector.
        For best results use a ColPali-native store like Qdrant with late interaction.
        """
        model = self._get_model()
        results: list[tuple[ColPaliPage, list[float]]] = []

        for page in pages:
            try:
                # Decode base64 back to PIL Image
                img_bytes = base64.b64decode(page.page_image_b64)
                pil_img   = Image.open(io.BytesIO(img_bytes)).convert("RGB")

                # Get ColPali embedding (returns tensor of shape [num_patches, dim])
                import torch
                with torch.no_grad():
                    embedding_tensor = model.encode_image(pil_img)

                # Average pool across patches → single vector
                if hasattr(embedding_tensor, "mean"):
                    avg_vector = embedding_tensor.mean(dim=0).tolist()
                else:
                    avg_vector = list(embedding_tensor)

                results.append((page, avg_vector))
            except Exception as e:
                print(f"ColPali embed failed for {page.source_file} page {page.page_idx}: {e}")

        return results

    def embed_query(self, query: str) -> list[float]:
        """
        Embed a text query for ColPali retrieval.
        ColPali uses a text encoder for queries (same late-interaction model).
        """
        model = self._get_model()
        import torch
        with torch.no_grad():
            embedding_tensor = model.encode_query(query)
        if hasattr(embedding_tensor, "mean"):
            return embedding_tensor.mean(dim=0).tolist()
        return list(embedding_tensor)
